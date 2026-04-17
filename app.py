import base64
import hashlib
import io
import json
import os
import re
import secrets
from datetime import datetime
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except Exception:
    PLOTLY_AVAILABLE = False

# =============================================================================
# Configuración general
# =============================================================================
APP_TITLE = "Plataforma de Evaluación de Personal - RR. HH. Guatemala"
DATA_DIR = os.path.dirname(os.path.abspath(__file__))
USERS_FILE = os.path.join(DATA_DIR, "usuarios_rrhh.json")
APP_VERSION = "1.0"

ROLE_OPTIONS = ["administrador", "reclutador", "consultor", "solo lectura"]
WORK_MODALITY_OPTIONS = ["presencial", "híbrido", "remoto"]
CONTRACT_TYPES = ["indefinido", "plazo fijo", "temporal", "servicios profesionales", "prácticas", "otro"]
EDUCATION_LEVELS = [
    "ninguno",
    "primaria",
    "básico",
    "diversificado",
    "técnico",
    "licenciatura",
    "maestría",
    "doctorado",
]

# =============================================================================
# Utilidades de persistencia y autenticación
# =============================================================================

def hash_password(password: str, salt: str | None = None) -> str:
    if salt is None:
        salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return f"{salt}${dk.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        salt, _ = stored.split("$", 1)
        return hash_password(password, salt) == stored
    except Exception:
        return False


def default_users() -> Dict[str, dict]:
    now = datetime.now().isoformat(timespec="seconds")
    return {
        "admin": {
            "full_name": "Administrador del sistema",
            "username": "admin",
            "password_hash": hash_password("admin123"),
            "role": "administrador",
            "active": True,
            "created_at": now,
            "updated_at": now,
        }
    }


def load_users() -> Dict[str, dict]:
    if not os.path.exists(USERS_FILE):
        users = default_users()
        save_users(users)
        return users
    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if "admin" not in data:
            data["admin"] = default_users()["admin"]
            save_users(data)
        return data
    except Exception:
        users = default_users()
        save_users(users)
        return users


def save_users(users: Dict[str, dict]) -> None:
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)


def init_state() -> None:
    if "users" not in st.session_state:
        st.session_state.users = load_users()
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "current_user" not in st.session_state:
        st.session_state.current_user = None
    if "vacancy" not in st.session_state:
        st.session_state.vacancy = default_vacancy()
    if "candidates" not in st.session_state:
        st.session_state.candidates = []
    if "analysis" not in st.session_state:
        st.session_state.analysis = pd.DataFrame()
    if "users_message" not in st.session_state:
        st.session_state.users_message = ""


def default_vacancy() -> dict:
    return {
        "nombre_puesto": "",
        "area": "",
        "nivel": "",
        "ubicacion": "",
        "modalidad": "presencial",
        "salario_ofrecido": 0.0,
        "tipo_contrato": "indefinido",
        "horario": "",
        "educacion_minima": "diversificado",
        "experiencia_minima_anios": 0.0,
        "competencias_tecnicas": [],
        "competencias_blandas": [],
        "idiomas": [],
        "viajar": False,
        "tolerancia_traslado": "media",
        "peso_cumplimiento_minimo": 15,
        "peso_tecnico": 20,
        "peso_experiencia": 15,
        "peso_salarial": 12,
        "peso_geografico": 10,
        "peso_disponibilidad": 8,
        "peso_blandas": 10,
        "peso_permanencia": 10,
        "peso_riesgo_rotacion": 10,
    }


# =============================================================================
# Normalización y soporte de texto
# =============================================================================

def normalize_text(value) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def split_items(value) -> List[str]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return []
    text = str(value).replace("\n", ",")
    parts = [normalize_text(x) for x in re.split(r"[,;/|]", text) if normalize_text(x)]
    return list(dict.fromkeys(parts))


def to_number(value, default=0.0) -> float:
    try:
        if value is None or (isinstance(value, str) and not value.strip()):
            return default
        if isinstance(value, str):
            value = value.replace(",", ".")
        return float(value)
    except Exception:
        return default


def education_index(level: str) -> int:
    level = normalize_text(level)
    try:
        return EDUCATION_LEVELS.index(level)
    except ValueError:
        return 0


def level_label(score: float) -> str:
    if score >= 80:
        return "recomendado"
    if score >= 60:
        return "evaluar en entrevista"
    return "no recomendado"


def semaforo(score: float) -> str:
    if score >= 80:
        return "🟢 Alto"
    if score >= 60:
        return "🟡 Medio"
    return "🔴 Bajo"


def safe_contains(container: List[str], item: str) -> bool:
    if not item:
        return False
    item = normalize_text(item)
    for c in container:
        c2 = normalize_text(c)
        if item == c2 or item in c2 or c2 in item:
            return True
    return False


def overlap_score(required: List[str], candidate: List[str]) -> Tuple[float, List[str], List[str]]:
    required_norm = [normalize_text(x) for x in required if normalize_text(x)]
    candidate_norm = [normalize_text(x) for x in candidate if normalize_text(x)]
    if not required_norm:
        return 100.0, [], candidate_norm
    matched = []
    missing = []
    for req in required_norm:
        if safe_contains(candidate_norm, req):
            matched.append(req)
        else:
            missing.append(req)
    return round((len(matched) / len(required_norm)) * 100, 2), matched, missing


# =============================================================================
# Evaluación y heurísticas
# =============================================================================

def parse_candidate(candidate: dict) -> dict:
    c = candidate.copy()
    c["habilidades_tecnicas_list"] = split_items(c.get("habilidades_tecnicas", []))
    c["habilidades_blandas_list"] = split_items(c.get("habilidades_blandas", []))
    c["idiomas_list"] = split_items(c.get("idiomas", []))
    c["certificaciones_list"] = split_items(c.get("certificaciones", []))
    c["formacion_list"] = split_items(c.get("formacion_academica", []))
    c["municipio"] = normalize_text(c.get("municipio", ""))
    c["departamento"] = normalize_text(c.get("departamento", ""))
    c["disponibilidad"] = normalize_text(c.get("disponibilidad", ""))
    c["interes_viajar"] = normalize_text(c.get("interes_viajar", ""))
    c["historial_permanencia"] = normalize_text(c.get("historial_permanencia", ""))
    c["expectativa_salarial"] = to_number(c.get("expectativa_salarial", 0))
    c["anos_experiencia"] = to_number(c.get("anos_experiencia", 0))
    c["promedio_permanencia_meses"] = to_number(c.get("promedio_permanencia_meses", np.nan), np.nan)
    c["cambios_laborales_5a"] = to_number(c.get("cambios_laborales_5a", np.nan), np.nan)
    c["nivel_educativo"] = normalize_text(c.get("nivel_educativo", ""))
    if not c["nivel_educativo"]:
        inferred = infer_education_level(c)
        c["nivel_educativo"] = inferred
    return c


def infer_education_level(candidate: dict) -> str:
    text = " ".join(
        [
            str(candidate.get("formacion_academica", "")),
            str(candidate.get("observaciones", "")),
            str(candidate.get("certificaciones", "")),
        ]
    ).lower()
    for level in reversed(EDUCATION_LEVELS):
        if level in text:
            return level
    return "diversificado"


def distance_factor(vacancy: dict, candidate: dict) -> float:
    vacancy_dept = normalize_text(vacancy.get("ubicacion", "")).split(",")[0].strip()
    candidate_dept = candidate.get("departamento", "")
    modality = normalize_text(vacancy.get("modalidad", "presencial"))
    tolerance = normalize_text(vacancy.get("tolerancia_traslado", "media"))
    travels = normalize_text(vacancy.get("viajar", False))
    travel_interest = candidate.get("interes_viajar", "")

    if not candidate_dept:
        base = 55
    elif vacancy_dept and candidate_dept == vacancy_dept:
        base = 100
    else:
        base = 72 if tolerance == "alta" else 58 if tolerance == "media" else 42

    if modality == "remoto":
        base = min(100, base + 18)
    elif modality == "híbrido":
        base = min(100, base + 8)

    if travels in {"sí", "si", "true", "1"} and any(x in travel_interest for x in ["sí", "si", "disponible", "siempre", "frecuente"]):
        base = min(100, base + 8)
    elif travels in {"sí", "si", "true", "1"} and travel_interest in {"", "no", "no disponible"}:
        base = max(0, base - 12)
    return float(np.clip(base, 0, 100))


def availability_score(vacancy: dict, candidate: dict) -> float:
    text = candidate.get("disponibilidad", "")
    modality = normalize_text(vacancy.get("modalidad", "presencial"))
    score = 60
    if any(k in text for k in ["inmediata", "inmediato", "full time", "tiempo completo", "disponible"]):
        score += 20
    if any(k in text for k in ["fines de semana", "rotativo", "turnos", "nocturno"]):
        score += 8
    if modality == "remoto" and any(k in text for k in ["teletrabajo", "remoto", "híbrido", "hibrido"]):
        score += 12
    if modality == "presencial" and any(k in text for k in ["presencial", "campo", "oficina"]):
        score += 10
    return float(np.clip(score, 0, 100))


def salary_fit_score(vacancy: dict, candidate: dict) -> float:
    offer = to_number(vacancy.get("salario_ofrecido", 0))
    expected = to_number(candidate.get("expectativa_salarial", 0))
    if offer <= 0:
        return 65.0
    if expected <= 0:
        return 70.0
    ratio = expected / offer
    if ratio <= 0.95:
        return 100.0
    if ratio <= 1.05:
        return 92.0
    if ratio <= 1.15:
        return 78.0
    if ratio <= 1.30:
        return 55.0
    if ratio <= 1.50:
        return 30.0
    return 10.0


def experience_score(vacancy: dict, candidate: dict) -> float:
    min_years = to_number(vacancy.get("experiencia_minima_anios", 0))
    years = to_number(candidate.get("anos_experiencia", 0))
    if min_years <= 0:
        return 100.0 if years > 0 else 65.0
    ratio = years / min_years if min_years else 1
    if ratio >= 1.5:
        return 100.0
    if ratio >= 1.0:
        return 90.0
    if ratio >= 0.75:
        return 75.0
    if ratio >= 0.5:
        return 58.0
    if ratio >= 0.25:
        return 35.0
    return 10.0


def technical_score(vacancy: dict, candidate: dict) -> Tuple[float, dict]:
    req_tech = vacancy.get("competencias_tecnicas", [])
    req_lang = vacancy.get("idiomas", [])
    cand_tech = candidate.get("habilidades_tecnicas_list", []) + candidate.get("certificaciones_list", [])
    cand_lang = candidate.get("idiomas_list", [])

    tech_score, matched_tech, missing_tech = overlap_score(req_tech, cand_tech)
    lang_score, matched_lang, missing_lang = overlap_score(req_lang, cand_lang)

    score = round((0.8 * tech_score) + (0.2 * lang_score), 2)
    details = {
        "matched_tech": matched_tech,
        "missing_tech": missing_tech,
        "matched_lang": matched_lang,
        "missing_lang": missing_lang,
        "tech_score": tech_score,
        "lang_score": lang_score,
    }
    return score, details


def soft_skills_score(vacancy: dict, candidate: dict) -> Tuple[float, dict]:
    req_soft = vacancy.get("competencias_blandas", [])
    cand_soft = candidate.get("habilidades_blandas_list", [])
    score, matched, missing = overlap_score(req_soft, cand_soft)
    return score, {"matched_soft": matched, "missing_soft": missing}


def minimum_requirement_score(vacancy: dict, candidate: dict, tech_details: dict) -> Tuple[float, List[str], List[str]]:
    notes_ok = []
    notes_bad = []
    vac_edu = education_index(vacancy.get("educacion_minima", "diversificado"))
    cand_edu = education_index(candidate.get("nivel_educativo", "diversificado"))
    if cand_edu >= vac_edu:
        notes_ok.append("cumple nivel educativo mínimo")
    else:
        notes_bad.append("no alcanza el nivel educativo mínimo")

    vac_exp = to_number(vacancy.get("experiencia_minima_anios", 0))
    cand_exp = to_number(candidate.get("anos_experiencia", 0))
    if cand_exp >= vac_exp:
        notes_ok.append("cumple experiencia mínima")
    else:
        notes_bad.append("experiencia inferior a la requerida")

    req_tech = [normalize_text(x) for x in vacancy.get("competencias_tecnicas", []) if normalize_text(x)]
    missing_tech = tech_details.get("missing_tech", [])
    if not req_tech:
        notes_ok.append("sin competencias técnicas obligatorias definidas")
    elif len(missing_tech) == 0:
        notes_ok.append("cumple competencias técnicas críticas")
    else:
        # Un faltante técnico crítico reduce el cumplimiento mínimo, pero no elimina la evaluación
        notes_bad.append(f"faltan {len(missing_tech)} competencias técnicas críticas")

    score = 100.0
    if cand_edu < vac_edu:
        score -= 45
    if cand_exp < vac_exp:
        score -= 35
    if req_tech and len(missing_tech) > 0:
        score -= min(35, len(missing_tech) * 12)
    return float(np.clip(score, 0, 100)), notes_ok, notes_bad


def permanence_estimate(vacancy: dict, candidate: dict, geo_score: float, salary_score: float, availability_score_value: float, experience_score_value: float, tech_score_value: float) -> Tuple[float, float, str, List[str]]:
    changes = candidate.get("cambios_laborales_5a", np.nan)
    avg_months = candidate.get("promedio_permanencia_meses", np.nan)
    if np.isnan(changes):
        changes = max(0, 5 - int(candidate.get("anos_experiencia", 0) // 2))
    if np.isnan(avg_months):
        avg_months = 24 if candidate.get("anos_experiencia", 0) >= 5 else 18 if candidate.get("anos_experiencia", 0) >= 3 else 12

    # Componentes explicables
    stability = float(np.clip((avg_months / 36) * 100, 0, 100))
    if changes == 0:
        churn_penalty = 0
    elif changes <= 1:
        churn_penalty = 4
    elif changes <= 3:
        churn_penalty = 12
    else:
        churn_penalty = 22
    stability = max(0, stability - churn_penalty)

    salary_component = salary_score
    geo_component = geo_score
    schedule_component = availability_score_value

    # Coherencia trayectoria: proxy de alineación técnica/experiencia
    coherence_component = round((0.55 * tech_score_value) + (0.45 * experience_score_value), 2)

    permanence = (
        0.35 * stability
        + 0.25 * salary_component
        + 0.20 * geo_component
        + 0.10 * schedule_component
        + 0.10 * coherence_component
    )
    permanence = float(np.clip(permanence, 0, 100))
    risk = float(np.clip(100 - permanence, 0, 100))

    explanation_parts = []
    explanation_parts.append(f"Estabilidad previa: {stability:.0f}/100")
    explanation_parts.append(f"Ajuste salarial: {salary_component:.0f}/100")
    explanation_parts.append(f"Factibilidad logística: {geo_component:.0f}/100")
    explanation_parts.append(f"Disponibilidad/jornada: {schedule_component:.0f}/100")
    explanation_parts.append(f"Coherencia trayectoria: {coherence_component:.0f}/100")
    explanation = "; ".join(explanation_parts)

    signals = []
    if stability >= 70:
        signals.append("historial laboral estable")
    if salary_component >= 85:
        signals.append("expectativa salarial alineada")
    if geo_component >= 75:
        signals.append("traslado manejable")
    if schedule_component >= 75:
        signals.append("jornada compatible")
    if coherence_component >= 75:
        signals.append("trayectoria coherente con la vacante")
    return permanence, risk, explanation, signals


def evaluate_candidate(vacancy: dict, candidate: dict) -> dict:
    c = parse_candidate(candidate)
    tech_score_value, tech_details = technical_score(vacancy, c)
    exp_score_value = experience_score(vacancy, c)
    salary_score_value = salary_fit_score(vacancy, c)
    geo_score_value = distance_factor(vacancy, c)
    availability_score_value = availability_score(vacancy, c)
    soft_score_value, soft_details = soft_skills_score(vacancy, c)
    minimum_score_value, min_ok, min_bad = minimum_requirement_score(vacancy, c, tech_details)
    permanence_value, risk_value, permanence_logic, permanence_signals = permanence_estimate(
        vacancy, c, geo_score_value, salary_score_value, availability_score_value, exp_score_value, tech_score_value
    )

    weights = {
        "minimum": to_number(vacancy.get("peso_cumplimiento_minimo", 15)),
        "technical": to_number(vacancy.get("peso_tecnico", 20)),
        "experience": to_number(vacancy.get("peso_experiencia", 15)),
        "salary": to_number(vacancy.get("peso_salarial", 12)),
        "geo": to_number(vacancy.get("peso_geografico", 10)),
        "availability": to_number(vacancy.get("peso_disponibilidad", 8)),
        "soft": to_number(vacancy.get("peso_blandas", 10)),
        "permanence": to_number(vacancy.get("peso_permanencia", 10)),
        "risk": to_number(vacancy.get("peso_riesgo_rotacion", 10)),
    }
    weight_sum = sum(weights.values())
    if weight_sum <= 0:
        weight_sum = 100

    # Normalizamos a 100 para que la suma sea consistente aunque el usuario cambie los pesos
    normalized = {k: (v / weight_sum) for k, v in weights.items()}
    final_score = (
        normalized["minimum"] * minimum_score_value
        + normalized["technical"] * tech_score_value
        + normalized["experience"] * exp_score_value
        + normalized["salary"] * salary_score_value
        + normalized["geo"] * geo_score_value
        + normalized["availability"] * availability_score_value
        + normalized["soft"] * soft_score_value
        + normalized["permanence"] * permanence_value
        + normalized["risk"] * (100 - risk_value)
    )
    final_score = float(np.clip(final_score, 0, 100))

    if minimum_score_value < 55:
        recommendation = "no recomendado"
    elif final_score >= 80 and permanence_value >= 65:
        recommendation = "recomendado"
    elif final_score >= 60:
        recommendation = "evaluar en entrevista"
    else:
        recommendation = "no recomendado"

    observations = []
    strengths = []
    weaknesses = []

    if tech_score_value >= 80:
        strengths.append("alto ajuste técnico")
    elif tech_score_value < 55:
        weaknesses.append("brecha técnica relevante")

    if exp_score_value >= 80:
        strengths.append("experiencia suficiente")
    elif exp_score_value < 55:
        weaknesses.append("experiencia por debajo de la expectativa")

    if salary_score_value >= 80:
        strengths.append("expectativa salarial compatible")
    elif salary_score_value < 55:
        weaknesses.append("expectativa salarial por encima de la oferta")

    if geo_score_value >= 80:
        strengths.append("traslado/geografía favorables")
    elif geo_score_value < 55:
        weaknesses.append("riesgo logístico por ubicación")

    if soft_score_value >= 75:
        strengths.append("competencias blandas alineadas")
    elif soft_score_value < 55:
        weaknesses.append("competencias blandas por reforzar")

    if permanence_value >= 70:
        strengths.append("alta probabilidad de permanencia")
    elif permanence_value < 55:
        weaknesses.append("riesgo de rotación elevado")

    if strengths:
        observations.append("Fortalezas: " + ", ".join(strengths) + ".")
    if weaknesses:
        observations.append("Alertas: " + ", ".join(weaknesses) + ".")
    if not observations:
        observations.append("Perfil equilibrado; se recomienda validar en entrevista técnica y de ajuste cultural.")

    if min_bad:
        observations.append("Revisión mínima: " + "; ".join(min_bad) + ".")

    explanation = (
        f"Puntaje final {final_score:.1f}/100. "
        f"Cumplimiento mínimo {minimum_score_value:.0f}/100; Técnico {tech_score_value:.0f}/100; Experiencia {exp_score_value:.0f}/100; "
        f"Salario {salary_score_value:.0f}/100; Geográfico {geo_score_value:.0f}/100; Disponibilidad {availability_score_value:.0f}/100; "
        f"Blandas {soft_score_value:.0f}/100; Permanencia {permanence_value:.0f}/100; Riesgo de rotación {risk_value:.0f}/100."
    )

    return {
        **c,
        "score_cumplimiento_minimo": round(minimum_score_value, 2),
        "score_tecnico": round(tech_score_value, 2),
        "score_experiencia": round(exp_score_value, 2),
        "score_salarial": round(salary_score_value, 2),
        "score_geografico": round(geo_score_value, 2),
        "score_disponibilidad": round(availability_score_value, 2),
        "score_blandas": round(soft_score_value, 2),
        "score_permanencia": round(permanence_value, 2),
        "score_riesgo_rotacion": round(risk_value, 2),
        "score_final": round(final_score, 2),
        "recomendacion": recommendation,
        "explicacion": explanation,
        "observaciones": " ".join(observations),
        "lógica_permanencia": permanence_logic,
        "indicadores_permanencia": ", ".join(permanence_signals) if permanence_signals else "sin señales adicionales",
        "cumple_minimos_ok": "; ".join(min_ok) if min_ok else "",
        "cumple_minimos_bad": "; ".join(min_bad) if min_bad else "",
        "tech_details": tech_details,
        "soft_details": soft_details,
    }


def evaluate_all(vacancy: dict, candidates: List[dict]) -> pd.DataFrame:
    rows = []
    for idx, cand in enumerate(candidates, start=1):
        result = evaluate_candidate(vacancy, cand)
        result["posible_id"] = cand.get("id_interno", f"C{idx:03d}")
        result["nombre"] = cand.get("nombre", f"Candidato {idx}")
        rows.append(result)
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df = df.sort_values(["score_final", "score_permanencia"], ascending=[False, False]).reset_index(drop=True)
    df.insert(0, "ranking", range(1, len(df) + 1))
    return df


# =============================================================================
# Plantillas y carga de candidatos
# =============================================================================

def candidate_template_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "nombre": "María López",
                "identificador_interno": "C-001",
                "formacion_academica": "Licenciatura en Administración",
                "nivel_educativo": "licenciatura",
                "experiencia_laboral": "3 años como analista administrativa",
                "anos_experiencia": 3,
                "habilidades_tecnicas": "Excel; SAP; reportes",
                "habilidades_blandas": "trabajo en equipo; comunicación; orden",
                "certificaciones": "Excel avanzado",
                "idiomas": "español; inglés",
                "municipio": "Guatemala",
                "departamento": "Guatemala",
                "expectativa_salarial": 6500,
                "disponibilidad": "inmediata",
                "interes_viajar": "sí",
                "historial_permanencia": "2 años promedio por empleo",
                "promedio_permanencia_meses": 24,
                "cambios_laborales_5a": 2,
                "observaciones": "",
            }
        ]
    )


def dataframe_to_template_bytes(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Candidatos")
        ws = writer.book["Candidatos"]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 35)
    return output.getvalue()


def parse_uploaded_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        df = pd.read_excel(uploaded_file)
    else:
        raise ValueError("Formato no soportado. Usa un archivo CSV o Excel (.xlsx/.xls).")
    df.columns = [normalize_text(c).replace(" ", "_") for c in df.columns]
    return df


def normalize_uploaded_candidates(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    aliases = {
        "nombre": ["nombre", "nombre_completo", "candidato"],
        "id_interno": ["id_interno", "identificador_interno", "codigo", "código"],
        "formacion_academica": ["formacion_academica", "formación_académica", "formacion", "formación"],
        "nivel_educativo": ["nivel_educativo", "educacion", "educación"],
        "experiencia_laboral": ["experiencia_laboral", "experiencia"],
        "anos_experiencia": ["anos_experiencia", "años_experiencia", "experiencia_anios", "experiencia_años"],
        "habilidades_tecnicas": ["habilidades_tecnicas", "competencias_tecnicas", "skills_tecnicos"],
        "habilidades_blandas": ["habilidades_blandas", "competencias_blandas"],
        "certificaciones": ["certificaciones"],
        "idiomas": ["idiomas"],
        "municipio": ["municipio"],
        "departamento": ["departamento", "depto"],
        "expectativa_salarial": ["expectativa_salarial", "salario_esperado", "pretension_salarial", "pretensión_salarial"],
        "disponibilidad": ["disponibilidad"],
        "interes_viajar": ["interes_viajar", "viajar"],
        "historial_permanencia": ["historial_permanencia", "estabilidad_laboral"],
        "promedio_permanencia_meses": ["promedio_permanencia_meses"],
        "cambios_laborales_5a": ["cambios_laborales_5a", "cambios_en_5_años", "cambios_en_5_anios"],
        "observaciones": ["observaciones", "notas"],
    }
    cols = {}
    lower_cols = {c.lower(): c for c in df.columns}
    for target, options in aliases.items():
        for opt in options:
            if opt.lower() in lower_cols:
                cols[target] = lower_cols[opt.lower()]
                break
    missing = [k for k in ["nombre"] if k not in cols]
    out = pd.DataFrame()
    for target, source in cols.items():
        out[target] = df[source]
    return out, missing


def add_candidates_from_df(df: pd.DataFrame) -> Tuple[int, List[str]]:
    added = 0
    warnings = []
    required = ["nombre"]
    for _, row in df.iterrows():
        candidate = row.to_dict()
        if not str(candidate.get("nombre", "")).strip():
            warnings.append("Se omitió una fila sin nombre.")
            continue
        if any(not str(candidate.get(req, "")).strip() for req in required):
            warnings.append(f"Candidato omitido por datos obligatorios faltantes: {candidate.get('nombre', 'sin nombre')}")
            continue
        if not str(candidate.get("id_interno", "")).strip():
            candidate["id_interno"] = f"C-{len(st.session_state.candidates) + added + 1:04d}"
        st.session_state.candidates.append(candidate)
        added += 1
    return added, warnings


# =============================================================================
# Excel profesional
# =============================================================================

def stylize_sheet(ws, header_row=1, freeze="A2"):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 42)


def build_excel_report(vacancy: dict, ranking_df: pd.DataFrame, candidates: List[dict]) -> bytes:
    output = io.BytesIO()
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen de la vacante"

    summary_rows = [
        ("Puesto", vacancy.get("nombre_puesto", "")),
        ("Área/Departamento", vacancy.get("area", "")),
        ("Nivel", vacancy.get("nivel", "")),
        ("Ubicación", vacancy.get("ubicacion", "")),
        ("Modalidad", vacancy.get("modalidad", "")),
        ("Salario ofrecido", f"Q {vacancy.get('salario_ofrecido', 0):,.2f}"),
        ("Tipo de contrato", vacancy.get("tipo_contrato", "")),
        ("Horario", vacancy.get("horario", "")),
        ("Educación mínima", vacancy.get("educacion_minima", "")),
        ("Experiencia mínima", f"{vacancy.get('experiencia_minima_anios', 0)} años"),
        ("Competencias técnicas", ", ".join(vacancy.get("competencias_tecnicas", []))),
        ("Competencias blandas", ", ".join(vacancy.get("competencias_blandas", []))),
        ("Idiomas", ", ".join(vacancy.get("idiomas", []))),
        ("Disponibilidad para viajar", "Sí" if vacancy.get("viajar") else "No"),
        ("Tolerancia al traslado", vacancy.get("tolerancia_traslado", "")),
    ]
    ws_summary["A1"] = "Resumen ejecutivo"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Campo"
    ws_summary["B3"] = "Valor"
    for cell in ws_summary[3]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    row = 4
    for k, v in summary_rows:
        ws_summary.cell(row=row, column=1, value=k)
        ws_summary.cell(row=row, column=2, value=v)
        row += 1
    stylize_sheet(ws_summary, header_row=3, freeze="A4")

    ws_rank = wb.create_sheet("Ranking general")
    if not ranking_df.empty:
        ranking_cols = [
            "ranking", "nombre", "posible_id", "score_final", "recomendacion",
            "score_tecnico", "score_experiencia", "score_salarial", "score_geografico",
            "score_disponibilidad", "score_blandas", "score_permanencia", "score_riesgo_rotacion",
            "observaciones"
        ]
        ranking_out = ranking_df[[c for c in ranking_cols if c in ranking_df.columns]].copy()
        for col in ranking_out.columns:
            if ranking_out[col].dtype != object:
                ranking_out[col] = ranking_out[col].round(2)
        for r in dataframe_to_excel_rows(ranking_out, ws_rank, start_row=1):
            pass
    else:
        ws_rank["A1"] = "Sin candidatos evaluados."
    stylize_sheet(ws_rank, header_row=1, freeze="A2")

    ws_detail = wb.create_sheet("Detalle por candidato")
    if not ranking_df.empty:
        detail_columns = [
            "ranking", "nombre", "posible_id", "score_final", "recomendacion",
            "score_cumplimiento_minimo", "score_tecnico", "score_experiencia",
            "score_salarial", "score_geografico", "score_disponibilidad",
            "score_blandas", "score_permanencia", "score_riesgo_rotacion",
            "explicacion", "observaciones", "lógica_permanencia", "indicadores_permanencia"
        ]
        detail_out = ranking_df[[c for c in detail_columns if c in ranking_df.columns]].copy()
        for col in detail_out.columns:
            if detail_out[col].dtype != object:
                detail_out[col] = detail_out[col].round(2)
        for r in dataframe_to_excel_rows(detail_out, ws_detail, start_row=1):
            pass
    else:
        ws_detail["A1"] = "Sin datos para detallar."
    stylize_sheet(ws_detail, header_row=1, freeze="A2")

    ws_obs = wb.create_sheet("Observaciones y recomendaciones")
    ws_obs.append(["Candidato", "Puntaje", "Recomendación", "Observaciones", "Explicación"])
    for cell in ws_obs[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for _, row in ranking_df.iterrows():
        ws_obs.append([
            row.get("nombre", ""),
            float(row.get("score_final", 0)),
            row.get("recomendacion", ""),
            row.get("observaciones", ""),
            row.get("explicacion", ""),
        ])
    stylize_sheet(ws_obs, header_row=1, freeze="A2")

    ws_params = wb.create_sheet("Parámetros de evaluación")
    params = [
        ["Parámetro", "Valor"],
        ["Peso cumplimiento mínimo", vacancy.get("peso_cumplimiento_minimo", 15)],
        ["Peso técnico", vacancy.get("peso_tecnico", 20)],
        ["Peso experiencia", vacancy.get("peso_experiencia", 15)],
        ["Peso salarial", vacancy.get("peso_salarial", 12)],
        ["Peso geográfico", vacancy.get("peso_geografico", 10)],
        ["Peso disponibilidad", vacancy.get("peso_disponibilidad", 8)],
        ["Peso blandas", vacancy.get("peso_blandas", 10)],
        ["Peso permanencia", vacancy.get("peso_permanencia", 10)],
        ["Peso riesgo rotación", vacancy.get("peso_riesgo_rotacion", 10)],
        ["Fecha de generación", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Total candidatos", len(candidates)],
    ]
    for row in params:
        ws_params.append(row)
    stylize_sheet(ws_params, header_row=1, freeze="A2")

    wb.save(output)
    return output.getvalue()


def dataframe_to_excel_rows(df: pd.DataFrame, ws, start_row=1):
    header = list(df.columns)
    ws.append(header)
    for cell in ws[start_row]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for _, row in df.iterrows():
        ws.append([row.get(col, "") for col in header])
    return True


# =============================================================================
# Interfaz de usuario
# =============================================================================

def login_view():
    st.title("Acceso seguro")
    st.caption("Sistema de evaluación de personal para Recursos Humanos en Guatemala")
    with st.form("login_form", clear_on_submit=False):
        username = st.text_input("Usuario")
        password = st.text_input("Contraseña", type="password")
        submitted = st.form_submit_button("Ingresar")
        if submitted:
            users = st.session_state.users
            user = users.get(username)
            if not user:
                st.error("Usuario no encontrado.")
            elif not user.get("active", True):
                st.error("El usuario está desactivado.")
            elif verify_password(password, user.get("password_hash", "")):
                st.session_state.authenticated = True
                st.session_state.current_user = {
                    "username": username,
                    "full_name": user.get("full_name", username),
                    "role": user.get("role", "solo lectura"),
                }
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")


def top_bar():
    user = st.session_state.current_user or {}
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        st.markdown(f"**Usuario:** {user.get('full_name', '-')}")
    with col2:
        st.markdown(f"**Rol:** {user.get('role', '-')}")
    with col3:
        if st.button("Salir", use_container_width=True):
            st.session_state.authenticated = False
            st.session_state.current_user = None
            st.rerun()


def role_allows_edit(role: str) -> bool:
    return role in {"administrador", "reclutador"}


def role_allows_view(role: str) -> bool:
    return role in {"administrador", "reclutador", "consultor", "solo lectura"}


def role_allows_export(role: str) -> bool:
    return role in {"administrador", "reclutador", "consultor"}


def sidebar_navigation() -> str:
    role = st.session_state.current_user.get("role", "solo lectura") if st.session_state.current_user else "solo lectura"
    pages = ["Inicio", "Vacante", "Candidatos", "Configuración", "Resultados", "Ranking", "Exportación"]
    if role == "administrador":
        pages.append("Administración")
    else:
        pages.append("Administración")
    # Solo administración visible para administradores
    if role != "administrador" and "Administración" in pages:
        pages = [p for p in pages if p != "Administración"]
    page = st.sidebar.radio("Navegación", pages)
    return page


def page_inicio():
    st.header("Inicio")
    st.write(
        "Esta plataforma evalúa candidatos de forma integral para vacantes en Guatemala, "
        "considerando ajuste técnico, salarial, geográfico y probabilidad de permanencia."
    )
    st.info(
        "Advertencia ética: esta herramienta solo apoya la decisión humana. No reemplaza el criterio profesional de RR. HH. "
        "y evita deliberadamente criterios discriminatorios o sensibles."
    )
    vacancy = st.session_state.vacancy
    if vacancy.get("nombre_puesto"):
        st.subheader("Vacante activa")
        st.write(f"**Puesto:** {vacancy.get('nombre_puesto')}\n\n**Ubicación:** {vacancy.get('ubicacion')}\n\n**Salario:** Q {vacancy.get('salario_ofrecido', 0):,.2f}")
    else:
        st.warning("Aún no has definido una vacante.")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Candidatos cargados", len(st.session_state.candidates))
    with col2:
        st.metric("Peso total configurado", f"{sum_weights(st.session_state.vacancy)}")
    with col3:
        st.metric("Versión", APP_VERSION)


def sum_weights(vacancy: dict) -> str:
    total = sum(
        [
            to_number(vacancy.get("peso_cumplimiento_minimo", 15)),
            to_number(vacancy.get("peso_tecnico", 20)),
            to_number(vacancy.get("peso_experiencia", 15)),
            to_number(vacancy.get("peso_salarial", 12)),
            to_number(vacancy.get("peso_geografico", 10)),
            to_number(vacancy.get("peso_disponibilidad", 8)),
            to_number(vacancy.get("peso_blandas", 10)),
            to_number(vacancy.get("peso_permanencia", 10)),
            to_number(vacancy.get("peso_riesgo_rotacion", 10)),
        ]
    )
    return f"{total:.0f}%"


def page_vacante():
    st.header("Vacante")
    st.caption("Define el puesto para que el sistema compare candidatos con base en una vacante real.")
    if not role_allows_edit(st.session_state.current_user.get("role", "")):
        st.warning("Tu rol tiene acceso de solo consulta. No puedes editar la vacante.")
    with st.form("vacancy_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            nombre_puesto = st.text_input("Nombre del puesto", value=st.session_state.vacancy.get("nombre_puesto", ""))
            area = st.text_input("Área o departamento", value=st.session_state.vacancy.get("area", ""))
            nivel_options = ["operativo", "asistente", "analista", "supervisor", "jefatura", "gerencia", "dirección"]
            current_nivel = st.session_state.vacancy.get("nivel", "analista")
            nivel_index = nivel_options.index(current_nivel) if current_nivel in nivel_options else 2
            nivel = st.selectbox("Nivel del cargo", nivel_options, index=nivel_index)
            ubicacion = st.text_input("Ubicación (municipio, departamento)", value=st.session_state.vacancy.get("ubicacion", ""))
            modalidad = st.selectbox("Modalidad de trabajo", WORK_MODALITY_OPTIONS, index=WORK_MODALITY_OPTIONS.index(st.session_state.vacancy.get("modalidad", "presencial")) if st.session_state.vacancy.get("modalidad", "presencial") in WORK_MODALITY_OPTIONS else 0)
        with c2:
            salario_ofrecido = st.number_input("Salario ofrecido en Q", min_value=0.0, step=100.0, value=float(st.session_state.vacancy.get("salario_ofrecido", 0.0)))
            tipo_contrato = st.selectbox("Tipo de contrato", CONTRACT_TYPES, index=CONTRACT_TYPES.index(st.session_state.vacancy.get("tipo_contrato", "indefinido")) if st.session_state.vacancy.get("tipo_contrato", "indefinido") in CONTRACT_TYPES else 0)
            horario = st.text_input("Horario", value=st.session_state.vacancy.get("horario", ""))
            educacion_minima = st.selectbox("Nivel educativo mínimo", EDUCATION_LEVELS, index=EDUCATION_LEVELS.index(st.session_state.vacancy.get("educacion_minima", "diversificado")) if st.session_state.vacancy.get("educacion_minima", "diversificado") in EDUCATION_LEVELS else 3)
        with c3:
            experiencia_minima_anios = st.number_input("Experiencia mínima (años)", min_value=0.0, step=0.5, value=float(st.session_state.vacancy.get("experiencia_minima_anios", 0.0)))
            viajar = st.checkbox("Requiere disponibilidad para viajar dentro del país", value=bool(st.session_state.vacancy.get("viajar", False)))
            tolerancia_traslado = st.select_slider("Tolerancia al traslado", options=["baja", "media", "alta"], value=st.session_state.vacancy.get("tolerancia_traslado", "media") if st.session_state.vacancy.get("tolerancia_traslado", "media") in ["baja", "media", "alta"] else "media")

        competencias_tecnicas = st.text_area("Competencias técnicas requeridas (separadas por coma)", value=", ".join(st.session_state.vacancy.get("competencias_tecnicas", [])))
        competencias_blandas = st.text_area("Competencias blandas requeridas (separadas por coma)", value=", ".join(st.session_state.vacancy.get("competencias_blandas", [])))
        idiomas = st.text_area("Idiomas requeridos (separados por coma)", value=", ".join(st.session_state.vacancy.get("idiomas", [])))

        st.subheader("Pesos de evaluación")
        w1, w2, w3 = st.columns(3)
        with w1:
            peso_cumplimiento_minimo = st.slider("Cumplimiento mínimo", 0, 30, int(st.session_state.vacancy.get("peso_cumplimiento_minimo", 15)))
            peso_tecnico = st.slider("Técnico", 0, 30, int(st.session_state.vacancy.get("peso_tecnico", 20)))
            peso_experiencia = st.slider("Experiencia", 0, 30, int(st.session_state.vacancy.get("peso_experiencia", 15)))
        with w2:
            peso_salarial = st.slider("Salario", 0, 30, int(st.session_state.vacancy.get("peso_salarial", 12)))
            peso_geografico = st.slider("Geográfico", 0, 30, int(st.session_state.vacancy.get("peso_geografico", 10)))
            peso_disponibilidad = st.slider("Disponibilidad", 0, 30, int(st.session_state.vacancy.get("peso_disponibilidad", 8)))
        with w3:
            peso_blandas = st.slider("Blandas", 0, 30, int(st.session_state.vacancy.get("peso_blandas", 10)))
            peso_permanencia = st.slider("Permanencia", 0, 30, int(st.session_state.vacancy.get("peso_permanencia", 10)))
            peso_riesgo_rotacion = st.slider("Riesgo rotación", 0, 30, int(st.session_state.vacancy.get("peso_riesgo_rotacion", 10)))

        submitted = st.form_submit_button("Guardar vacante")
        if submitted:
            if not role_allows_edit(st.session_state.current_user.get("role", "")):
                st.error("No tienes permisos para modificar esta información.")
            else:
                st.session_state.vacancy = {
                    "nombre_puesto": nombre_puesto.strip(),
                    "area": area.strip(),
                    "nivel": nivel,
                    "ubicacion": ubicacion.strip(),
                    "modalidad": modalidad,
                    "salario_ofrecido": salario_ofrecido,
                    "tipo_contrato": tipo_contrato,
                    "horario": horario.strip(),
                    "educacion_minima": educacion_minima,
                    "experiencia_minima_anios": experiencia_minima_anios,
                    "competencias_tecnicas": split_items(competencias_tecnicas),
                    "competencias_blandas": split_items(competencias_blandas),
                    "idiomas": split_items(idiomas),
                    "viajar": viajar,
                    "tolerancia_traslado": tolerancia_traslado,
                    "peso_cumplimiento_minimo": peso_cumplimiento_minimo,
                    "peso_tecnico": peso_tecnico,
                    "peso_experiencia": peso_experiencia,
                    "peso_salarial": peso_salarial,
                    "peso_geografico": peso_geografico,
                    "peso_disponibilidad": peso_disponibilidad,
                    "peso_blandas": peso_blandas,
                    "peso_permanencia": peso_permanencia,
                    "peso_riesgo_rotacion": peso_riesgo_rotacion,
                }
                st.success("Vacante guardada correctamente.")
    total = sum_weights(st.session_state.vacancy)
    if total != "100%":
        st.warning(f"Los pesos actuales suman {total}. La app los normaliza automáticamente para el cálculo final.")


def manual_candidate_form() -> dict:
    with st.form("manual_candidate_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            nombre = st.text_input("Nombre completo")
            id_interno = st.text_input("Identificador interno opcional")
            formacion_academica = st.text_input("Formación académica")
            nivel_educativo = st.selectbox("Nivel educativo", EDUCATION_LEVELS, index=3)
        with c2:
            experiencia_laboral = st.text_area("Experiencia laboral resumida")
            anos_experiencia = st.number_input("Años de experiencia", min_value=0.0, step=0.5, value=0.0)
            habilidades_tecnicas = st.text_area("Habilidades técnicas (coma)")
            habilidades_blandas = st.text_area("Habilidades blandas (coma)")
        with c3:
            certificaciones = st.text_area("Certificaciones (coma)")
            idiomas = st.text_area("Idiomas (coma)")
            municipio = st.text_input("Municipio")
            departamento = st.text_input("Departamento")
        c4, c5, c6 = st.columns(3)
        with c4:
            expectativa_salarial = st.number_input("Expectativa salarial en Q", min_value=0.0, step=100.0, value=0.0)
            disponibilidad = st.text_input("Disponibilidad")
        with c5:
            interes_viajar = st.selectbox("Interés en viajar", ["", "sí", "no"], index=0)
            promedio_permanencia_meses = st.number_input("Promedio de permanencia (meses)", min_value=0.0, step=1.0, value=0.0)
        with c6:
            cambios_laborales_5a = st.number_input("Cambios laborales últimos 5 años", min_value=0.0, step=1.0, value=0.0)
        historial_permanencia = st.text_area("Historial de permanencia en empleos anteriores")
        observaciones = st.text_area("Observaciones")
        submitted = st.form_submit_button("Agregar candidato")
        if submitted:
            if not role_allows_edit(st.session_state.current_user.get("role", "")):
                st.error("Tu rol no permite registrar candidatos.")
                return {}
            if not nombre.strip():
                st.error("El nombre es obligatorio.")
                return {}
            return {
                "nombre": nombre.strip(),
                "id_interno": id_interno.strip(),
                "formacion_academica": formacion_academica.strip(),
                "nivel_educativo": nivel_educativo,
                "experiencia_laboral": experiencia_laboral.strip(),
                "anos_experiencia": anos_experiencia,
                "habilidades_tecnicas": split_items(habilidades_tecnicas),
                "habilidades_blandas": split_items(habilidades_blandas),
                "certificaciones": split_items(certificaciones),
                "idiomas": split_items(idiomas),
                "municipio": municipio.strip(),
                "departamento": departamento.strip(),
                "expectativa_salarial": expectativa_salarial,
                "disponibilidad": disponibilidad.strip(),
                "interes_viajar": interes_viajar,
                "promedio_permanencia_meses": promedio_permanencia_meses,
                "cambios_laborales_5a": cambios_laborales_5a,
                "historial_permanencia": historial_permanencia.strip(),
                "observaciones": observaciones.strip(),
            }
    return {}


def page_candidatos():
    st.header("Candidatos")
    st.caption("Registra personas manualmente o importa un archivo Excel/CSV con candidatos.")
    if not role_allows_edit(st.session_state.current_user.get("role", "")):
        st.warning("Tu rol no permite agregar ni importar candidatos.")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Carga manual")
        new_candidate = manual_candidate_form()
        if new_candidate:
            if not new_candidate.get("id_interno"):
                new_candidate["id_interno"] = f"C-{len(st.session_state.candidates) + 1:04d}"
            st.session_state.candidates.append(new_candidate)
            st.success(f"Candidato agregado: {new_candidate['nombre']}")
    with col2:
        st.subheader("Importación masiva")
        template_bytes = dataframe_to_template_bytes(candidate_template_df())
        st.download_button(
            "Descargar plantilla Excel",
            data=template_bytes,
            file_name="plantilla_candidatos_rrhh.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        uploaded = st.file_uploader("Subir Excel o CSV", type=["xlsx", "xls", "csv"])
        if uploaded is not None:
            if not role_allows_edit(st.session_state.current_user.get("role", "")):
                st.error("Tu rol no permite cargar archivos.")
            else:
                try:
                    raw_df = parse_uploaded_file(uploaded)
                    normalized_df, missing = normalize_uploaded_candidates(raw_df)
                    if missing:
                        st.error("Falta al menos la columna obligatoria: nombre.")
                    else:
                        st.dataframe(normalized_df.head(10), use_container_width=True)
                        if st.button("Importar archivo", use_container_width=True):
                            added, warnings = add_candidates_from_df(normalized_df)
                            st.success(f"Se importaron {added} candidatos.")
                            for w in warnings[:5]:
                                st.warning(w)
                except Exception as e:
                    st.error(f"No fue posible leer el archivo: {e}")

    st.divider()
    st.subheader("Candidatos registrados")
    if st.session_state.candidates:
        df = pd.DataFrame(st.session_state.candidates)
        show_cols = [c for c in ["id_interno", "nombre", "departamento", "municipio", "nivel_educativo", "anos_experiencia", "expectativa_salarial"] if c in df.columns]
        st.dataframe(df[show_cols], use_container_width=True, hide_index=True)
        if role_allows_edit(st.session_state.current_user.get("role", "")):
            if st.button("Limpiar todos los candidatos", type="secondary"):
                st.session_state.candidates = []
                st.success("Listado de candidatos limpiado.")
    else:
        st.info("Aún no hay candidatos registrados.")


def page_configuracion():
    st.header("Configuración")
    st.caption("Ajustes operativos para la evaluación y el contexto guatemalteco.")
    st.write("La lógica evita criterios discriminatorios y se enfoca en variables laborales objetivas.")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Factores considerados**")
        st.write(
            "- Salario en quetzales\n"
            "- Municipio y departamento\n"
            "- Facilidad de traslado\n"
            "- Disponibilidad para viajar\n"
            "- Modalidad presencial/híbrida/remota\n"
            "- Nivel educativo\n"
            "- Experiencia formal e informal\n"
            "- Idiomas\n"
            "- Estabilidad laboral previa\n"
            "- Expectativa salarial vs oferta"
        )
    with c2:
        st.markdown("**Permisos por rol**")
        st.write(
            "- **administrador**: acceso total y gestión de usuarios.\n"
            "- **reclutador**: puede registrar vacantes y candidatos.\n"
            "- **consultor**: lectura y exportación.\n"
            "- **solo lectura**: consulta sin edición."
        )

    st.subheader("Advertencia ética")
    st.warning(
        "Este sistema solo apoya la decisión. Debe evitarse cualquier uso para discriminar por sexo, religión, estado civil, "
        "etnia, orientación política, apariencia física u otros atributos sensibles."
    )


def filter_rank_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    st.sidebar.subheader("Filtros de ranking")
    min_score = st.sidebar.slider("Puntaje mínimo", 0, 100, 0)
    max_score = st.sidebar.slider("Puntaje máximo", 0, 100, 100)
    options_location = ["Todas"] + sorted([x for x in df.get("departamento", pd.Series(dtype=str)).fillna("").astype(str).unique().tolist() if x])
    location = st.sidebar.selectbox("Departamento", options_location)
    options_edu = ["Todos"] + sorted([x for x in df.get("nivel_educativo", pd.Series(dtype=str)).fillna("").astype(str).unique().tolist() if x])
    edu = st.sidebar.selectbox("Nivel educativo", options_edu)
    options_av = ["Todos"] + sorted([x for x in df.get("disponibilidad", pd.Series(dtype=str)).fillna("").astype(str).unique().tolist() if x])
    avail = st.sidebar.selectbox("Disponibilidad", options_av)
    salary_max = int(df["expectativa_salarial"].fillna(0).max()) if "expectativa_salarial" in df.columns and not df.empty else 0
    salary_limit = st.sidebar.slider("Expectativa salarial máxima (Q)", 0, max(1000, salary_max), max(1000, salary_max))

    filtered = df[(df["score_final"] >= min_score) & (df["score_final"] <= max_score)]
    if location != "Todas" and "departamento" in filtered.columns:
        filtered = filtered[filtered["departamento"].fillna("").astype(str) == location]
    if edu != "Todos" and "nivel_educativo" in filtered.columns:
        filtered = filtered[filtered["nivel_educativo"].fillna("").astype(str) == edu]
    if avail != "Todos" and "disponibilidad" in filtered.columns:
        filtered = filtered[filtered["disponibilidad"].fillna("").astype(str) == avail]
    if "expectativa_salarial" in filtered.columns:
        filtered = filtered[filtered["expectativa_salarial"].fillna(0) <= salary_limit]
    return filtered


def page_resultados():
    st.header("Resultados")
    if not st.session_state.candidates:
        st.info("Debes cargar candidatos para ver resultados.")
        return
    ranking_df = evaluate_all(st.session_state.vacancy, st.session_state.candidates)
    st.session_state.analysis = ranking_df.copy()
    if ranking_df.empty:
        st.info("No hay resultados disponibles.")
        return
    filtered = filter_rank_df(ranking_df)
    st.subheader("Tabla interactiva")
    st.dataframe(
        filtered[
            [
                "ranking", "nombre", "score_final", "recomendacion", "score_tecnico",
                "score_experiencia", "score_salarial", "score_geografico", "score_disponibilidad",
                "score_blandas", "score_permanencia", "score_riesgo_rotacion", "observaciones"
            ]
        ],
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Indicadores principales")
    top = ranking_df.iloc[0]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Mejor puntaje", f"{top['score_final']:.1f}")
    c2.metric("Recomendación", top["recomendacion"])
    c3.metric("Permanencia estimada", f"{top['score_permanencia']:.1f}")
    c4.metric("Riesgo de rotación", f"{top['score_riesgo_rotacion']:.1f}")

    st.subheader("Gráficos comparativos")
    if PLOTLY_AVAILABLE:
        chart_df = filtered.copy() if not filtered.empty else ranking_df.copy()
        fig = px.bar(
            chart_df,
            x="nombre",
            y="score_final",
            color="recomendacion",
            title="Ranking por puntaje total",
        )
        fig.update_layout(xaxis_title="Candidato", yaxis_title="Puntaje total")
        st.plotly_chart(fig, use_container_width=True)

        fig2 = px.scatter(
            chart_df,
            x="score_tecnico",
            y="score_permanencia",
            size="score_final",
            color="recomendacion",
            hover_name="nombre",
            title="Ajuste técnico vs permanencia estimada",
        )
        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.line_chart(ranking_df.set_index("nombre")["score_final"])

    st.subheader("Detalle por candidato")
    for _, row in ranking_df.iterrows():
        with st.expander(f"#{int(row['ranking'])} {row['nombre']} - {row['score_final']:.1f} ({row['recomendacion']})"):
            st.write(row["explicacion"])
            st.write(f"**Semáforo:** {semaforo(row['score_final'])}")
            st.write(f"**Observaciones:** {row['observaciones']}")
            st.write(f"**Lógica de permanencia:** {row['lógica_permanencia']}")
            st.write(f"**Señales de permanencia:** {row['indicadores_permanencia']}")


def page_ranking():
    st.header("Ranking")
    if st.session_state.analysis.empty:
        ranking_df = evaluate_all(st.session_state.vacancy, st.session_state.candidates)
    else:
        ranking_df = st.session_state.analysis.copy()
    if ranking_df.empty:
        st.info("No hay datos para ranking.")
        return
    st.success("Ranking calculado automáticamente de mayor a menor puntaje final.")
    cols = [c for c in ["ranking", "nombre", "score_final", "score_tecnico", "score_experiencia", "score_salarial", "score_geografico", "score_permanencia", "score_riesgo_rotacion", "recomendacion", "observaciones"] if c in ranking_df.columns]
    st.dataframe(ranking_df[cols], use_container_width=True, hide_index=True)
    if PLOTLY_AVAILABLE:
        fig = px.line(ranking_df, x="ranking", y="score_final", markers=True, title="Curva de ranking")
        st.plotly_chart(fig, use_container_width=True)
    st.caption("La posición se define por el puntaje final; a igualdad de puntaje, se prioriza mayor permanencia estimada.")


def page_exportacion():
    st.header("Exportación")
    if st.session_state.analysis.empty:
        ranking_df = evaluate_all(st.session_state.vacancy, st.session_state.candidates)
    else:
        ranking_df = st.session_state.analysis.copy()
    if ranking_df.empty:
        st.info("No hay datos para exportar.")
        return
    if not role_allows_export(st.session_state.current_user.get("role", "")):
        st.warning("Tu rol no tiene permiso de exportación.")
        return
    report_bytes = build_excel_report(st.session_state.vacancy, ranking_df, st.session_state.candidates)
    st.download_button(
        "Descargar informe Excel",
        data=report_bytes,
        file_name=f"informe_rrhh_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.write("El archivo incluye hojas con resumen ejecutivo, ranking, detalle por candidato, observaciones y parámetros.")


def page_admin():
    st.header("Administración")
    if st.session_state.current_user.get("role") != "administrador":
        st.error("Acceso restringido. Solo el administrador puede administrar usuarios.")
        return

    users = st.session_state.users
    st.subheader("Registrar nuevo usuario")
    with st.form("new_user_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            username = st.text_input("Usuario")
            full_name = st.text_input("Nombre completo")
        with c2:
            password = st.text_input("Contraseña", type="password")
            role = st.selectbox("Rol", ROLE_OPTIONS, index=1)
        with c3:
            active = st.checkbox("Activo", value=True)
        submitted = st.form_submit_button("Crear usuario")
        if submitted:
            if not username.strip() or not password.strip():
                st.error("Usuario y contraseña son obligatorios.")
            elif username in users:
                st.error("El usuario ya existe.")
            else:
                now = datetime.now().isoformat(timespec="seconds")
                users[username] = {
                    "full_name": full_name.strip() or username,
                    "username": username,
                    "password_hash": hash_password(password),
                    "role": role,
                    "active": active,
                    "created_at": now,
                    "updated_at": now,
                }
                save_users(users)
                st.session_state.users = users
                st.success("Usuario creado correctamente.")

    st.subheader("Usuarios registrados")
    user_rows = []
    for u in users.values():
        user_rows.append(
            {
                "username": u.get("username"),
                "full_name": u.get("full_name"),
                "role": u.get("role"),
                "active": u.get("active", True),
                "created_at": u.get("created_at"),
                "updated_at": u.get("updated_at"),
            }
        )
    user_df = pd.DataFrame(user_rows)
    if user_df.empty:
        st.info("No hay usuarios registrados.")
        return
    edited = st.data_editor(
        user_df,
        use_container_width=True,
        hide_index=True,
        disabled=["username", "created_at"],
        column_config={
            "role": st.column_config.SelectboxColumn("Rol", options=ROLE_OPTIONS),
            "active": st.column_config.CheckboxColumn("Activo"),
        },
        key="users_editor",
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        save_changes = st.button("Guardar cambios", use_container_width=True)
    with c2:
        st.write("")
    with c3:
        st.write("")

    if save_changes:
        updated_users = users.copy()
        for _, row in edited.iterrows():
            username = row["username"]
            if username not in updated_users:
                continue
            updated_users[username]["full_name"] = row["full_name"]
            updated_users[username]["role"] = row["role"]
            updated_users[username]["active"] = bool(row["active"])
            updated_users[username]["updated_at"] = datetime.now().isoformat(timespec="seconds")
        save_users(updated_users)
        st.session_state.users = updated_users
        st.success("Usuarios actualizados.")

    st.subheader("Desactivar usuario")
    if len(user_rows) > 1:
        usernames = [u["username"] for u in user_rows if u["username"] != "admin"]
        target = st.selectbox("Selecciona usuario", usernames)
        if st.button("Desactivar seleccionado"):
            if target in users:
                users[target]["active"] = False
                users[target]["updated_at"] = datetime.now().isoformat(timespec="seconds")
                save_users(users)
                st.session_state.users = users
                st.success(f"Usuario {target} desactivado.")

    st.info("El administrador es el único perfil autorizado para crear y modificar usuarios.")


# =============================================================================
# Aplicación principal
# =============================================================================

def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    init_state()
    st.title(APP_TITLE)
    st.caption("Evaluación integral de candidatos para vacantes en Guatemala")

    if not st.session_state.authenticated:
        login_view()
        return

    top_bar()
    page = sidebar_navigation()
    role = st.session_state.current_user.get("role", "solo lectura")

    if page == "Inicio":
        page_inicio()
    elif page == "Vacante":
        if role in {"administrador", "reclutador", "consultor", "solo lectura"}:
            page_vacante()
    elif page == "Candidatos":
        page_candidatos()
    elif page == "Configuración":
        page_configuracion()
    elif page == "Resultados":
        page_resultados()
    elif page == "Ranking":
        page_ranking()
    elif page == "Exportación":
        page_exportacion()
    elif page == "Administración":
        page_admin()

    st.sidebar.divider()
    st.sidebar.caption("Uso interno de RR. HH. No sustituye la decisión profesional.")


if __name__ == "__main__":
    main()

"""
Explicación breve de funcionamiento:
- El puntaje final se calcula con una suma ponderada normalizada de cumplimiento mínimo, ajuste técnico, experiencia, salario, ubicación, disponibilidad, competencias blandas, permanencia estimada y riesgo de rotación.
- El ranking ordena a los candidatos por score_final de mayor a menor; en empate, se prioriza la permanencia estimada.
- La permanencia se estima con una heurística transparente basada en estabilidad previa, ajuste salarial, logística, jornada y coherencia de trayectoria.
- El reporte Excel se genera con pandas y openpyxl en varias hojas: resumen de vacante, ranking general, detalle por candidato, observaciones y parámetros de evaluación, con formato profesional.
- El acceso de administrador funciona con sesión en Streamlit; el usuario admin / admin123 se crea automáticamente en la primera ejecución y solo el administrador puede registrar, editar o desactivar usuarios.
"""
